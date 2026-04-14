from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
from collections import defaultdict
from datetime import datetime, timezone
from typing import TYPE_CHECKING, Any, Callable, Optional

if TYPE_CHECKING:
    from .indexing import BaseMetadataIndex, BaseSupervisionStore, BaseVectorIndex

logger = logging.getLogger(__name__)

def _parse_ts(value: str) -> Optional[datetime]:
    """Parse an ISO 8601 timestamp string. Returns None on empty/malformed."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None

# -----------------------------------------------------------------------------
# Public entry point
# -----------------------------------------------------------------------------

async def ingest_ticket(
    ticket_key: str,
    jira_client: Any,
    llm_client: Optional[Any] = None,
    embedding_client: Optional[Any] = None,
    dict_path: Optional[str] = None,
    config: Optional[Any] = None,
    coarse_index: Optional[Any] = None,
    fine_index: Optional[Any] = None,
    metadata_index: Optional[Any] = None,
    supervision_store: Optional[Any] = None,
    force_reprocess: bool = False,
) -> dict:
    """Fetch, process, and index a single Jira ticket using multi-document ingestion."""
    from jira_ingestion.config import JiraIngestionConfig
    from ingestion.indexing import index_retrieval_view, index_supervision_view

    cfg = config if config is not None else JiraIngestionConfig()
    ticket_data = await jira_client.get_ticket_data(ticket_key, config=cfg)

    if coarse_index is not None and not force_reprocess:
        existing_record = coarse_index.get(ticket_key)
        if isinstance(existing_record, dict):
            existing_updated_raw = (existing_record.get("metadata") or {}).get("updated_at") or existing_record.get("updated_at") or ""
            incoming_updated_raw = (ticket_data.get("fields") or {}).get("updated") or ""
            existing_updated = _parse_ts(str(existing_updated_raw)) if existing_updated_raw else None
            incoming_updated = _parse_ts(str(incoming_updated_raw)) if incoming_updated_raw else None
            if existing_updated is not None and incoming_updated is not None and incoming_updated <= existing_updated:
                return existing_record

    document = await ingest_ticket_payload(
        ticket_data=ticket_data,
        jira_client=jira_client,
        llm_client=llm_client,
        embedding_client=embedding_client,
        dict_path=dict_path,
        config=cfg,
    )

    if coarse_index is not None and fine_index is not None and metadata_index is not None:
        index_retrieval_view(document, coarse_index, fine_index, metadata_index)

    if supervision_store is not None:
        index_supervision_view(document, supervision_store)

    return document

async def ingest_ticket_payload(
    ticket_data: dict,
    jira_client: Any,
    llm_client: Optional[Any] = None,
    embedding_client: Optional[Any] = None,
    dict_path: Optional[str] = None,
    config: Optional[Any] = None,
) -> dict:
    """Process an already fetched Jira ticket payload into a unified document."""
    from jira_ingestion.config import JiraIngestionConfig

    cfg = config if config is not None else JiraIngestionConfig()

    attachments = ticket_data.get("attachments", []) or []

    # Pre-download viable attachments. Unlike the previous top-1 gate, this stage
    # can fetch all bounded extractable attachments so downstream assembly is not
    # constrained by a filename-score cutoff.
    prefetched = await _prefetch_attachment_bytes(jira_client, attachments, cfg)

    def cached_download(att: dict) -> bytes:
        att_id = str(att.get("id") or att.get("filename") or "")
        cached = prefetched.get(att_id)
        if cached is None:
            raise RuntimeError(
                f"Bytes not prefetched for attachment '{att.get('filename')}'. "
                "Increase max_prefetch_attachments or relax the prefetch filter."
            )
        return cached

    document = assemble_document(
        ticket_data=ticket_data,
        download_fn=cached_download,
        llm_client=llm_client,
        embedding_client=embedding_client,
        dict_path=dict_path,
        config=cfg,
    )

    return document

# -----------------------------------------------------------------------------
# Core assembly
# -----------------------------------------------------------------------------

def assemble_document(
    ticket_data: dict,
    download_fn: Optional[Callable[[dict], bytes]] = None,
    llm_client: Optional[Any] = None,
    embedding_client: Optional[Any] = None,
    dict_path: Optional[str] = None,
    config: Optional[Any] = None,
) -> Optional[dict]:
    """
    Assemble a unified v4.0 document from raw ticket data.

    Output layers:
      raw         - immutable source data
      observed    - retrieval view (no supervision labels)
      supervision - ground-truth labels
      derived     - LLM outputs
    """
    from jira_ingestion.config import JiraIngestionConfig
    from description import build_description_chunks, classify_description
    from embedding import embed_batch
    from entities import ensure_default_dictionaries, extract_entities, load_entity_dictionaries
    from metadata import (
        classify_links,
        extract_adf_text,
        extract_comments_enriched,
        extract_metadata,
        extract_product_fields,
        extract_stage_labels,
    )
    from summary import extract_chunk_keywords, generate_derived_artifacts, generate_summary
    from ingestion.attachment_routing import route_attachments, get_routing_candidates
    from ingestion.value_stream_mapping_service import resolve_value_stream_mapping

    cfg = config if config is not None else JiraIngestionConfig()
    resolved_dict_path = dict_path or getattr(cfg, "entity_dict_path", None)

    ticket_key = str(ticket_data.get("key", ""))
    fields: dict = ticket_data.get("fields", {})
    now = datetime.now(timezone.utc).isoformat()

    # 1) Metadata extraction (retrieval-safe fields only)
    meta = extract_metadata(fields, ticket_key, config=cfg)
    classified_links = classify_links(fields.get("issuelinks", []))
    meta["classified_links"] = classified_links
    jira_field_map: dict = getattr(cfg, "jira_field_map", {}) or {}

    # 2) Product / supervision labels (not retrieval metadata)
    product_fields = extract_product_fields(fields, jira_field_map)
    product_stage_labels = extract_stage_labels(fields, jira_field_map)

    # 3) Comments - enriched (raw + cleaned)
    comments_enriched = extract_comments_enriched(fields.get("comment") or {})

    # 4) Description
    raw_description_value = fields.get("description")
    if isinstance(raw_description_value, (dict, list)):
        raw_description = extract_adf_text(raw_description_value)
    else:
        raw_description = raw_description_value or ""
    desc_class, desc_data = classify_description(raw_description)
    description_cleaned = desc_data["text"] if desc_data else ""

    # 5) Attachment triage - routing, not one-doc gatekeeping
    attachments = ticket_data.get("attachments", []) or []
    primary, supporting, att_quality, routing_artifact = route_attachments(
        attachments=attachments,
        ticket_summary=meta.get("summary") or ticket_key,
        download_fn=download_fn,
        enable_image_ocr=bool(getattr(cfg, "enable_standalone_image_ocr", True)),
    )

    chunk_candidates = get_routing_candidates(routing_artifact)
    logger.info(
        "Routing primary: %s, supporting: %d, quality: %s, candidates: %d",
        primary.get("filename") if primary else None,
        len(supporting),
        att_quality,
        len(chunk_candidates),
    )

    # 6) Multi-document extraction + chunking
    processed_attachments = _extract_chunk_candidates(chunk_candidates, download_fn, cfg)
    processed_by_id = {a["attachment_id"]: a for a in processed_attachments}
    primary_attachment_id = str(primary.get("id") or primary.get("filename") or "") if primary else ""
    primary_processed = processed_by_id.get(primary_attachment_id)
    primary_attachment_text = primary_processed.get("text", "") if primary_processed else ""
    primary_attachment_name = primary_processed.get("filename", "") if primary_processed else ""

    # 7) Attachment inventory (accurate extraction state)
    attachment_inventory = _build_attachment_inventory(
        attachments=attachments,
        triage_artifact=routing_artifact,
        processed_attachments=processed_attachments,
    )

    # 8) Attachment chunks - all viable docs, weighted by role
    attachment_chunks: list[dict] = []
    attachment_texts_by_role: dict[str, list[str]] = defaultdict(list)
    for doc in processed_attachments:
        attachment_chunks.extend(doc["chunks"])
        if doc.get("text"):
            attachment_texts_by_role[doc.get("doc_role") or "supporting_doc"].append(doc["text"])

    # 9) Description chunks
    has_attachment = bool(processed_attachments)
    desc_chunks = build_description_chunks(desc_class, desc_data, meta, has_attachment)
    for chunk in desc_chunks:
        chunk["attachment_id"] = ""
        chunk["attachment_name"] = ""
        chunk["no_attachment"] = True
        chunk.setdefault("section_title", "Description")
        chunk.setdefault("chunk_granularity", "description")
        chunk.setdefault("weight_multiplier", 0.70 if has_attachment else 0.85)
        chunk.setdefault("extraction_confidence", 0.90)

    # 10) Comment chunks (lower weight - supplementary evidence)
    comment_limit = int(getattr(cfg, "max_comment_chunks", 2) or 2)
    comment_max_chars = int(getattr(cfg, "comment_chunk_chars", 1500) or 1500)
    comment_chunks: list[dict] = []
    for i, comment_text in enumerate((comments_enriched.get("comments_cleaned") or [])[:comment_limit]):
        if not comment_text:
            continue
        comment_chunks.append(
            {
                "chunk_id": f"comment-{i}",
                "source": "comment",
                "section_title": f"Comment {i + 1}",
                "text": comment_text[:comment_max_chars],
                "word_count": len(comment_text.split()),
                "is_boilerplate": False,
                "weight_multiplier": 0.45,
                "extraction_confidence": 0.55,
                "attachment_id": "",
                "attachment_name": "",
                "no_attachment": True,
                "chunk_granularity": "comment",
            }
        )

    # 10a) Description-linked attachment references (e.g., SharePoint idea card links)
    description_link_chunks = _build_description_link_chunks(ticket_data.get("description_attachments") or [])

    raw_chunks = attachment_chunks + desc_chunks + comment_chunks + description_link_chunks

    # 11) Quality tier
    content_source = _derive_content_source(processed_attachments, desc_class, comments_enriched)
    quality_tier = _determine_quality_tier_safe(content_source, desc_class, raw_chunks, routing_artifact)

    # 12) Section chunks (attachment-aware and non-duplicative)
    section_chunks = _build_attachment_aware_section_chunks(processed_attachments, cfg)

    ticket_source_url = _build_ticket_source_url(ticket_data, ticket_key)
    _attach_chunk_identity(
        raw_chunks,
        ticket_key=ticket_key,
        source_url=ticket_source_url,
        default_attachment_id=primary_attachment_id,
        default_attachment_name=primary_attachment_name,
    )
    _attach_chunk_identity(
        section_chunks,
        ticket_key=ticket_key,
        source_url=ticket_source_url,
        default_attachment_id=primary_attachment_id,
        default_attachment_name=primary_attachment_name,
    )

    # 13) Retrieval chunks - balanced evidence pool
    retrieval_chunks = _build_retrieval_chunks(
        raw_chunks=raw_chunks,
        section_chunks=section_chunks,
        section_only_chunks=bool(getattr(cfg, "section_only_chunks", False)),
        include_section_rollups=bool(getattr(cfg, "include_section_rollups_in_retrieval", False)),
    )
    retrieval_chunks = _split_oversized_chunks(
        retrieval_chunks,
        max_tokens=int(getattr(cfg, "section_max_tokens", 700) or 700),
        overlap_tokens=int(getattr(cfg, "section_overlap_tokens", 80) or 80),
    )
    _attach_chunk_identity(
        chunks=retrieval_chunks,
        ticket_key=ticket_key,
        source_url=ticket_source_url,
        default_attachment_id=primary_attachment_id,
        default_attachment_name=primary_attachment_name,
    )

    # 14) Summary
    all_for_summary = retrieval_chunks
    if getattr(cfg, "skip_llm_summary", False):
        logger.info("Skipping LLM summary generation (skip_llm_summary=True)")
        summary_text = _heuristic_summary(meta, description_cleaned, processed_attachments)
    else:
        try:
            summary_text = generate_summary(
                quality_tier=quality_tier,
                chunks=all_for_summary,
                metadata=meta,
                llm_client=llm_client,
                model=getattr(cfg, "llm_model", None),
            )
        except Exception as exc:
            logger.warning("LLM summary generation failed: %s", exc)
            summary_text = _heuristic_summary(meta, description_cleaned, processed_attachments)

    # 15) Entity extraction
    ensure_default_dictionaries(resolved_dict_path)
    dictionaries = load_entity_dictionaries(resolved_dict_path)
    entity_mentions = extract_entities(retrieval_chunks, meta, dictionaries)

    # 16) Retrieval views (multiple focused representations)
    retrieval_views = []
    if getattr(cfg, "enable_retrieval_views", True):
        retrieval_views = _build_retrieval_views(
            meta=meta,
            description_cleaned=description_cleaned,
            primary_attachment_text=primary_attachment_text,
            supporting_texts=attachment_texts_by_role,
            comments_cleaned=comments_enriched.get("comments_cleaned", []),
            chunks=retrieval_chunks,
        )

    retrieval_text = _build_retrieval_text(
        summary_str=summary_text,
        description_cleaned=description_cleaned,
        primary_attachment_text=primary_attachment_text,
        supporting_texts=attachment_texts_by_role,
        retrieval_views=retrieval_views,
    )

    # 17) Embeddings
    texts_to_embed: list[str] = [summary_text]
    texts_to_embed.extend(c.get("text", "") for c in retrieval_chunks)
    texts_to_embed.append(meta.get("metadata_text", ""))

    if embedding_client is not None:
        try:
            embeddings = embed_batch(texts_to_embed, embedding_client, model=getattr(cfg, "embedding_model", None))
        except Exception as exc:
            logger.warning("Embedding generation failed: %s", exc)
            embeddings = [[] for _ in texts_to_embed]
    else:
        logger.warning("No embedding client - embeddings will be empty lists")
        embeddings = [[] for _ in texts_to_embed]

    idx = 0
    summary_embedding = embeddings[idx] if idx < len(embeddings) else []
    idx += 1
    for chunk in retrieval_chunks:
        chunk["embedding"] = embeddings[idx] if idx < len(embeddings) else []
        idx += 1
    metadata_embedding = embeddings[idx] if idx < len(embeddings) else []

    # 18) Context keywords per chunk (LLM)
    if getattr(cfg, "skip_llm_keywords", False):
        logger.info("Skipping LLM keyword extraction (skip_llm_keywords=True)")
    else:
        try:
            extract_chunk_keywords(
                retrieval_chunks,
                llm_client=llm_client,
                model=getattr(cfg, "llm_model", None),
            )
        except Exception as exc:
            logger.warning("Chunk keyword extraction failed: %s", exc)

    # 19) Traceability and provenance
    vs_mapping = resolve_value_stream_mapping(ticket_data, classified_links)

    # 20) Value stream with epics mapping
    from ingestion.value_stream_mapping_service import resolve_value_stream_epics_mapping
    vs_epics_mapping = resolve_value_stream_epics_mapping(ticket_data, classified_links)
    has_gold_vs = bool(vs_mapping.get("vs_links", []))
    has_gold_products = bool(product_fields.get("impacted_products") or ({}).get("names"))
    avg_confidence = sum(float(c.get("extraction_confidence", 0.0) or 0.0) for c in retrieval_chunks) / max(len(retrieval_chunks), 1)
    source_quality_score = _source_quality_score(quality_tier, avg_confidence, processed_attachments)

    provenance = {
        "quality_tier": quality_tier,
        "has_description": desc_class not in ("empty", "junk"),
        "attachment_count_total": len(attachments),
        "attachment_count_viable": len(routing_artifact.get("chunk_candidates", [])),
        "attachment_count_extracted": len(processed_attachments),
        "has_primary_attachment": primary is not None,
        "primary_evidence_type": primary.get("doc_role") if primary else "none",
        "selection_reason": routing_artifact.get("selection_reason", ""),
        "content_source": content_source,
        "source_quality_score": round(source_quality_score, 4),
        "processed_attachment_ids": [a.get("attachment_id") for a in processed_attachments],
        "contributing_attachment_ids": sorted({c.get("attachment_id") for c in retrieval_chunks if c.get("attachment_id")}),
    }

    # 21) Stats
    raw_attachment_desc = sum(1 for c in retrieval_chunks if c.get("source") == "description")
    doc_chunk_count = sum(1 for c in retrieval_chunks if c.get("attachment_id"))
    section_chunk_count = sum(1 for c in section_chunks)
    stats = {
        "chunk_count": len(retrieval_chunks),
        "doc_chunk_count": doc_chunk_count,
        "description_chunk_count": raw_attachment_desc,
        "comment_chunk_count": len(comment_chunks),
        "section_count": section_chunk_count,
        "word_count_total": sum(c.get("word_count", 0) for c in retrieval_chunks),
        "attachment_inventory_count": len(attachment_inventory),
        "entity_mention_count": sum(len(v) for v in entity_mentions.values()),
        "substantive_comment_count": comments_enriched.get("substantive_count", 0),
    }

    # 22) Derived layer - LLM structured outputs + entity signal fallback
    if getattr(cfg, "skip_llm_derived", False):
        logger.info("Skipping LLM derived artifacts (skip_llm_derived=True)")
        derived = {
            "ticket_summary_llm": "",
            "problem_summary_llm": "",
            "solution_summary_llm": "",
            "capability_keywords_llm": [],
            "product_mentions_llm": [],
            "business_entities_llm": [],
        }
    else:
        try:
            derived = generate_derived_artifacts(
                quality_tier=quality_tier,
                chunks=all_for_summary,
                metadata=meta,
                llm_client=llm_client,
                model=getattr(cfg, "llm_model", None),
            )
        except Exception as exc:
            logger.warning("LLM derived artifacts failed: %s", exc)
            derived = {
                "ticket_summary_llm": "",
                "problem_summary_llm": "",
                "solution_summary_llm": "",
                "capability_keywords_llm": [],
                "product_mentions_llm": [],
                "business_entities_llm": [],
            }

    # Always keep a summary available
    if not derived.get("ticket_summary_llm") and summary_text:
        derived["ticket_summary_llm"] = summary_text

    # Entity-signal fallbacks
    if not derived.get("capability_keywords_llm"):
        derived["capability_keywords_llm"] = [
            m.get("term") if isinstance(m, dict) else m
            for m in entity_mentions.get("capabilities", [])
        ]
    if not derived.get("product_mentions_llm"):
        derived["product_mentions_llm"] = [
            m.get("term") if isinstance(m, dict) else m
            for m in entity_mentions.get("products", [])
        ]
    if not derived.get("business_entities_llm"):
        ent_list = []
        if meta.get("business_unit"):
            ent_list.append(meta["business_unit"])
        if meta.get("product_area"):
            ent_list.append(meta["product_area"])
        ent_list.extend(meta.get("components", []))
        derived["business_entities_llm"] = ent_list[:8]

    # 23) Value stream supervision labels - structured
    vs_links = list(vs_mapping.get("vs_links", []))
    vs_names = list(vs_mapping.get("vs_names", []))
    vs_ids = list(vs_mapping.get("vs_ids", []))
    vs_statuses = list(vs_mapping.get("vs_statuses", []))
    linked_value_streams = [
        {
            "id": link.get("id", ""),
            "name": link.get("name", ""),
            "status": link.get("status", ""),
            "summary_raw": link.get("summary_raw", link.get("name", "")),
        }
        for link in (vs_mapping.get("linked_value_streams", []))
    ]

    # Denormalize ticket-level value-stream mapping onto chunk records for exports/search indexes.
    for chunk in retrieval_chunks:
        chunk["mapped_value_stream_ids"] = list(vs_ids)
        chunk["mapped_value_stream_names"] = list(vs_names)

    # 24) Assemble final document (v4.0)
    return {
        "ticket_key": ticket_key,
        "schema_version": "4.0",
        "ingested_at": now,
        # RAW LAYER
        "raw": {
            "description": raw_description,
            "comments": comments_enriched,
            "issue_links": fields.get("issuelinks", []),
            "theme": ticket_data.get("theme", {}),
            "epics": ticket_data.get("epics", []),
            "attachment_inventory": attachment_inventory,
        },
        # OBSERVED (RETRIEVAL) LAYER
        "observed": {
            "quality_tier": quality_tier,
            "created_at": meta.get("created", ""),
            "updated_at": fields.get("updated", ""),
            "content_source": content_source,
            "provenance": provenance,
            "triage": routing_artifact,
            "description_class": desc_class,
            "description_cleaned": description_cleaned,
            "primary_attachment_text": primary_attachment_text[:5000] if primary_attachment_text else "",
            "comments_cleaned": comments_enriched.get("comments_cleaned", []),
            "retrieval_views": retrieval_views,
            "retrieval_text": retrieval_text,
            "summary_text": summary_text,
            "summary_embedding": summary_embedding,
            "metadata": meta,
            "metadata_text": meta.get("metadata_text", ""),
            "metadata_embedding": metadata_embedding,
            "stats": stats,
            "chunks": retrieval_chunks,
            "section_chunks": section_chunks,
            "raw_chunks": raw_chunks,
            "entity_mentions": entity_mentions,
            "processed_attachments": processed_attachments,
        },
        # SUPERVISION LAYER
        "supervision": {
            "vs_labels": vs_names,
            "vs_label_source": str(vs_mapping.get("label_source", "jira_issuelinks")),
            "linked_value_stream_ids": vs_ids,
            "linked_value_stream_names": vs_names,
            "linked_value_stream_statuses": vs_statuses,
            "linked_value_streams": linked_value_streams,
            "theme_links": vs_links,
            "value_stream_epics": vs_epics_mapping.get("vs_with_epics", []),
            "all_epics": vs_epics_mapping.get("epics", []),
            "epic_ids": vs_epics_mapping.get("all_epic_ids", []),
            "epic_names": vs_epics_mapping.get("all_epic_names", []),
            "has_primary_attachment": primary is not None,
            "impacted_products": product_fields.get("impacted_products"),
            "impacted_it_products": product_fields.get("impacted_it_products"),
            "product_stage_labels": product_stage_labels,
            "trainability": {
                "has_gold_vs": has_gold_vs,
                "has_gold_products": has_gold_products,
                "is_trainable_for_vs": has_gold_vs and quality_tier in ("A", "B", "C"),
                "is_trainable_for_product": has_gold_products and quality_tier in ("A", "B", "C"),
                "source_quality_score": round(source_quality_score, 4),
                "label_snapshot_time": now,
            },
        },
        # DERIVED LAYER
        "derived": derived,
    }

# -----------------------------------------------------------------------------
# Attachment processing
# -----------------------------------------------------------------------------

EXTRACTABLE_PREFETCH_EXTS = {"pptx", "ppt", "pdf", "docx", "doc", "xlsx", "xls", "csv", "png", "jpg", "jpeg", "gif", "bmp", "tiff", "webp"}

async def _prefetch_attachment_bytes(jira_client: Any, attachments: list[dict], cfg: Any) -> dict[str, bytes]:
    """Prefetch bounded extractable attachments. This is no longer top-1 only."""
    max_prefetch = int(getattr(cfg, "max_prefetch_attachments", 5) or 5)
    max_size = int(getattr(cfg, "max_prefetch_attachment_size", 50_000_000) or 50_000_000)

    selected: list[dict] = []
    enable_image_ocr = bool(getattr(cfg, "enable_standalone_image_ocr", True))
    for att in attachments:
        if len(selected) >= max_prefetch:
            break
        filename = (att.get("filename") or "").lower()
        ext = filename.rsplit(".", 1)[-1] if "." in filename else ""
        size = int(att.get("size", 0) or 0)
        if ext not in EXTRACTABLE_PREFETCH_EXTS:
            continue
        if not enable_image_ocr and ext in ("png", "jpg", "jpeg", "gif", "bmp", "tiff", "webp"):
            continue
        if size and size < max_size:
            selected.append(att)
        elif not size:
            selected.append(att)

    prefetched: dict[str, bytes] = {}
    for att in selected:
        att_id = str(att.get("id") or att.get("filename") or "")
        try:
            prefetched[att_id] = await jira_client.download_attachment(att)
        except Exception as exc:
            logger.warning("Prefetch failed for %s: %s", att.get("filename"), exc)
    return prefetched

def _extract_chunk_candidates(
    chunk_candidates: list[dict],
    download_fn: Optional[Callable[[dict], bytes]],
    cfg: Any,
) -> list[dict]:
    processed: list[dict] = []
    if download_fn is None:
        return processed

    max_chunk_attachments = int(getattr(cfg, "max_chunk_attachments", 6) or 6)
    for att in chunk_candidates[:max_chunk_attachments]:
        try:
            file_bytes = download_fn(att)
        except Exception as exc:
            logger.warning("Download failed for %s: %s", att.get("filename"), exc)
            processed.append(_failed_attachment_record(att, str(exc)))
            continue

        try:
            processed.append(_extract_attachment(file_bytes, att, cfg))
        except Exception as exc:
            logger.warning("Attachment extraction failed for %s: %s", att.get("filename"), exc)
            processed.append(_failed_attachment_record(att, str(exc)))

    return processed

def _extract_attachment(file_bytes: bytes, att: dict, cfg: Any) -> dict:
    """Extract a viable attachment into normalized chunks with doc-aware metadata."""
    att_id = str(att.get("id") or _ext_from_name(att.get("filename", "")) or "")
    filename = str(att.get("filename") or att_id)
    doc_role = str(att.get("doc_role") or "supporting_doc")
    triage_score = float(att.get("triage_score", 0) or 0)
    att_scores = att.get("att_scores") or {}
    confidence = float(att_scores.get("extraction_quality", 0.55) or 0.55)
    retrieval_weight = _doc_role_weight(doc_role) * max(0.35, float(att_scores.get("retrieval_readiness", 0.5) or 0.5))

    # Prefer triage extraction result if it already exists and includes chunks.
    triage_extracted = att.get("extracted") or {}
    result = triage_extracted if triage_extracted.get("chunks") else {}
    if not result:
        ext = att.get("ext", "").lower()
        if ext in ("pptx", "ppt"):
            from .extraction.pptx import extract_pptx
            result = extract_pptx(file_bytes, max_slides=int(getattr(cfg, "max_slides", 60) or 60))
        elif ext == "pdf":
            from .extraction.pdf import extract_pdf
            ocr_enabled=bool(getattr(cfg, "ocr_enabled", False))
            max_pages=int(getattr(cfg, "max_slides", 60) or 60)
            result = extract_pdf(file_bytes, ocr_enabled=ocr_enabled, max_pages=max_pages)
        elif ext in ("docx", "doc"):
            from .extraction.docx import extract_docx
            result = extract_docx(file_bytes)
        elif ext == "csv":
            result = _extract_csv(file_bytes)
        elif ext in ("xlsx", "xls"):
            result = _extract_xlsx(file_bytes)
        elif ext in ("png", "jpg", "jpeg", "gif", "bmp", "tiff", "webp") and bool(getattr(cfg, "enable_standalone_image_ocr", True)):
            result = _extract_image(file_bytes, filename)
        else:
            result = {"chunks": []}

    normalized_chunks = _normalize_attachment_chunks(
        result=result,
        att=att,
        doc_role=doc_role,
        retrieval_weight=retrieval_weight,
        extraction_confidence=confidence,
    )

    text = "\n".join(c.get("text", "") for c in normalized_chunks if c.get("text") and not c.get("is_boilerplate"))
    text = text[:int(getattr(cfg, "max_attachment_text_chars", 200_000) or 200_000)]

    preview = text[:500]
    return {
        "attachment_id": att_id,
        "filename": filename,
        "doc_role": doc_role,
        "ext": att.get("ext", ""),
        "size": att.get("size", 0),
        "triage_score": triage_score,
        "triage_reasons": list(att.get("triage_reasons", [])),
        "att_scores": att_scores,
        "confirmed": bool(att.get("confirmed", False)),
        "extraction_status": "extracted" if normalized_chunks else "failed",
        "extraction_method": _extract_method_for_ext(att.get("ext", "")),
        "extraction_confidence": confidence,
        "text": text,
        "preview": preview,
        "chunks": normalized_chunks,
    }

def _extract_csv(file_bytes: bytes) -> dict:
    try:
        text = file_bytes.decode("utf-8", errors="ignore")
    except Exception:
        text = file_bytes.decode("latin-1", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"chunks": []}

    headers = rows[0]
    chunks: list[dict] = []
    for idx, row in enumerate(rows[1:101], start=1):
        pairs = []
        for h, v in zip(headers, row):
            if v:
                pairs.append(f"{h}: {v}")
        row_text = " | ".join(pairs).strip()
        if not row_text:
            continue
        chunks.append(
            {
                "chunk_id": f"csv-row-{idx}",
                "source": "csv_row",
                "section_title": f"CSV Row {idx}",
                "text": row_text,
                "row_num": idx,
                "is_boilerplate": False,
            }
        )
    return {"chunks": chunks}

def _extract_xlsx(file_bytes: bytes) -> dict:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        logger.warning("openpyxl unavailable for XLSX extraction: %s", exc)
        return {"chunks": []}

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("XLSX extraction failed to open workbook: %s", exc)
        return {"chunks": []}

    chunks: list[dict] = []
    for ws in wb.worksheets[:5]:
        rows = list(ws.iter_rows(values_only=True, max_row=50))
        if not rows:
            continue
        headers = [str(c) if c is not None else f"Col_{i+1}" for i, c in enumerate(rows[0])]
        for idx, row in enumerate(rows[1:51], start=1):
            pairs = []
            for h, v in zip(headers, row):
                if v is not None and v != "":
                    pairs.append(f"{h}: {v}")
            row_text = " | ".join(pairs).strip()
            if not row_text:
                continue
            chunks.append(
                {
                    "chunk_id": f"sheet-{ws.title}-row-{idx}",
                    "source": "sheet_row",
                    "section_title": f"{ws.title} Row {idx}",
                    "sheet_name": ws.title,
                    "row_num": idx,
                    "text": row_text,
                    "is_boilerplate": False,
                }
            )
    return {"chunks": chunks}

def _extract_image(file_bytes: bytes, filename: str) -> dict:
    from .extraction.markitdown import extract_image_text

    text = extract_image_text(file_bytes, filename=filename)
    if not text:
        return {"chunks": []}

    return {
        "chunks": [
            {
                "chunk_id": "image-ocr-1",
                "source": "image_ocr",
                "section_title": "Image OCR",
                "text": text,
                "is_boilerplate": False,
                "word_count": len(text.split()),
            }
        ]
    }

def _normalize_attachment_chunks(
    result: dict,
    att: dict,
    doc_role: str,
    retrieval_weight: float,
    extraction_confidence: float,
) -> list[dict]:
    att_id = str(att.get("id") or att.get("filename") or "")
    filename = str(att.get("filename") or att_id)
    ext = str(att.get("ext") or _ext_from_name(filename) or "")
    source_default = _source_for_ext(ext)

    chunks: list[dict] = []
    for idx, raw in enumerate(result.get("chunks", []) or []):
        text = raw.get("text") or ""
        if not str(text).strip():
            continue

        chunk = dict(raw)
        chunk["chunk_id"] = str(chunk.get("chunk_id") or f"{att_id}-{idx}")
        chunk["source"] = str(chunk.get("source") or source_default)
        chunk["source_format"] = ext
        chunk["attachment_id"] = att_id
        chunk["attachment_name"] = filename
        chunk["doc_role"] = doc_role
        chunk["word_count"] = int(chunk.get("word_count") or len(str(text).split()))
        chunk["weight_multiplier"] = round(float(chunk.get("weight_multiplier") or retrieval_weight), 4)
        chunk["extraction_confidence"] = round(float(chunk.get("extraction_confidence") or extraction_confidence), 4)
        chunk["chunk_granularity"] = str(chunk.get("chunk_granularity") or _granularity_for_source(chunk["source"]))
        chunk["is_boilerplate"] = bool(chunk.get("is_boilerplate", False))
        chunk.setdefault("section_title", chunk.get("slide_title") or chunk.get("page_title") or filename)
        chunks.append(chunk)

    return chunks

def _failed_attachment_record(att: dict, error: str) -> dict:
    att_id = str(att.get("id") or att.get("filename") or "")
    return {
        "attachment_id": att_id,
        "filename": str(att.get("filename") or att_id),
        "doc_role": str(att.get("doc_role") or "supporting_doc"),
        "ext": str(att.get("ext") or _ext_from_name(att.get("filename", "")) or ""),
        "size": int(att.get("size") or 0),
        "triage_score": float(att.get("triage_score", 0) or 0),
        "triage_reasons": list(att.get("triage_reasons", [])),
        "scores": att.get("att_scores") or {},
        "confirmed": bool(att.get("confirmed", False)),
        "extraction_status": "failed",
        "extraction_method": "none",
        "extraction_confidence": 0.0,
        "text": "",
        "preview": "",
        "chunks": [],
        "error": error,
    }

def _build_attachment_inventory(
    attachments: list[dict],
    triage_artifact: dict,
    processed_attachments: list[dict],
) -> list[dict]:
    processed_by_id = {a["attachment_id"]: a for a in processed_attachments}
    primary_id = str(triage_artifact.get("primary_attachment_id") or "")
    supporting_ids = set([str(x) for x in triage_artifact.get("supporting_attachment_ids", [])])
    excluded_ids = set([str(x) for x in triage_artifact.get("excluded_attachment_ids", [])])
    score_by_id = {
        str(x.get("attachment_id", "")): x for x in triage_artifact.get("per_attachment_scores", [])
    }

    inventory: list[dict] = []
    for att in attachments:
        att_id = str(att.get("id") or att.get("filename") or "")
        score_rec = score_by_id.get(att_id) or {}
        processed = processed_by_id.get(att_id)

        if att_id in excluded_ids:
            status = "filtered"
        elif processed and processed.get("extraction_status") == "extracted":
            status = "extracted"
        elif processed and processed.get("extraction_status") == "failed":
            status = "failed"
        elif att_id in supporting_ids:
            status = "supporting"
        elif att_id == primary_id:
            status = "primary"
        else:
            status = "skipped"

        inventory.append(
            {
                "attachment_id": att_id,
                "filename": att.get("filename", ""),
                "mime_type": att.get("mimeType", ""),
                "size": att.get("size", 0),
                "created_at": att.get("created", ""),
                "is_primary": att_id == primary_id,
                "is_supporting": att_id in supporting_ids,
                "doc_role": score_rec.get("doc_role") or (processed.get("doc_role") if processed else "none"),
                "extraction_status": status,
                "extraction_method": processed.get("extraction_method", "none") if processed else "none",
                "extraction_confidence": processed.get("extraction_confidence", 0.0) if processed else 0.0,
                "triage_score": score_rec.get("triage_score", 0),
                "triage_reasons": score_rec.get("triage_reasons", []),
                "scores": score_rec.get("scores", {}),
                "raw_text": processed.get("text", "") if processed else "",
                "cleaned_text": processed.get("text", "") if processed else "",
                "text_preview": processed.get("preview", "") if processed else "",
                "error": processed.get("error") if processed else None,
            }
        )

    return inventory

def _build_description_link_chunks(description_attachments: list[dict]) -> list[dict]:
    chunks: list[dict] = []
    for idx, att in enumerate(description_attachments):
        filename = str(att.get("filename") or "")
        url = str(att.get("content") or "")
        display_text = str(att.get("display_text") or filename or "Attachment Link")
        if not url or not filename:
            continue

        text = (
            f"Description includes linked attachment: [{display_text}].\n"
            f"Filename: {filename}, URL: {url}"
        ).strip()

        chunks.append(
            {
                "chunk_id": f"description-link-{idx}",
                "source": "description_attachment_link",
                "section_title": "Description Attachment Link",
                "text": text,
                "word_count": len(text.split()),
                "is_boilerplate": False,
                "weight_multiplier": 0.35,
                "extraction_confidence": 0.50,
                "attachment_id": str(att.get("id") or ""),
                "attachment_name": filename,
                "chunk_granularity": "link_reference",
            }
        )

    return chunks

# -----------------------------------------------------------------------------
# Retrieval construction
# -----------------------------------------------------------------------------

_PROBLEM_PATTERNS = re.compile(
    r"\b(problem|challenge|pain|issue|gap|opportunity|need|current state|why now)\b",
    re.IGNORECASE,
)
_SOLUTION_PATTERNS = re.compile(
    r"\b(solution|approach|proposal|capability|what we.?.?re building|recommendation|initiative|scope|deliverable)\b",
    re.IGNORECASE,
)
_VALUE_PATTERNS = re.compile(
    r"\b(value|benefit|roi|metric|kpi|success criteria|outcome|impact|saving|cost|revenue|efficiency)\b",
    re.IGNORECASE,
)

def _build_attachment_aware_section_chunks(processed_attachments: list[dict], cfg: Any) -> list[dict]:
    """Build section rollups per attachment instead of blindly duplicating the primary doc."""
    section_chunks: list[dict] = []

    try:
        from chunking import build_section_chunks as external_build_section_chunks
    except Exception:
        external_build_section_chunks = None

    for doc in processed_attachments:
        # Build section chunks primarily for slide/page style content.
        base_chunks = [
            c
            for c in doc.get("chunks", [])
            if c.get("source") in ("pptx_slide", "pdf_page", "doc_section")
            and not c.get("is_boilerplate")
        ]

        if len(base_chunks) < int(getattr(cfg, "section_min_slides", 3) or 3):
            continue

        built: list[dict] = []
        if external_build_section_chunks is not None:
            try:
                built = external_build_section_chunks(base_chunks) or []
            except Exception as exc:
                built = []

        if not built:
            built = _fallback_build_section_chunks(base_chunks, doc)

        for chunk in built:
            chunk["attachment_id"] = doc["attachment_id"]
            chunk["attachment_name"] = doc["filename"]
            chunk["doc_role"] = doc["doc_role"]
            chunk["source"] = "section"
            chunk["source_format"] = doc.get("ext") or _ext_from_name(doc["filename"])
            chunk["chunk_granularity"] = "section_rollup"
            chunk["weight_multiplier"] = round(doc_role_weight(doc["doc_role"]) * 0.75, 4)
            chunk["extraction_confidence"] = round(doc.get("extraction_confidence", 0.6), 4)
            chunk["word_count"] = len(chunk.get("text", "").split())
            chunk.setdefault("is_boilerplate", False)
            section_chunks.append(chunk)

    return section_chunks

def _fallback_build_section_chunks(base_chunks: list[dict], doc: dict) -> list[dict]:
    groups: dict[str, list[str]] = defaultdict(list)
    for chunk in base_chunks:
        title = str(chunk.get("section_title") or chunk.get("slide_title") or doc.get("filename") or "Attachment").strip()
        groups[title].append(chunk.get("text", ""))

    out: list[dict] = []
    for idx, (title, texts) in enumerate(groups.items(), start=1):
        merged = "\n".join(t for t in texts if t).strip()
        if not merged:
            continue
        out.append(
            {
                "chunk_id": f"section-{doc.get('attachment_id')}-{idx}",
                "section_title": title,
                "text": merged[:4000],
            }
        )

    return out

def _build_retrieval_views(
    meta: dict,
    description_cleaned: str,
    primary_attachment_text: str,
    supporting_texts: dict[str, list[str]],
    comments_cleaned: list[str],
    chunks: list[dict],
) -> list[dict]:
    """Build multiple focused retrieval views from balanced ticket content."""
    summary = meta.get("summary", "")
    components = ", ".join(meta.get("components", []))
    labels = ", ".join(meta.get("labels", []))
    business_unit = meta.get("business_unit", "")
    product_area = meta.get("product_area", "")

    overview_parts = [summary]
    if business_unit:
        overview_parts.append(f"Business Unit: {business_unit}")
    if product_area:
        overview_parts.append(f"Product Area: {product_area}")
    if components:
        overview_parts.append(f"Components: {components}")
    if labels:
        overview_parts.append(f"Labels: {labels}")
    if description_cleaned:
        overview_parts.append(description_cleaned[:800])
    if primary_attachment_text:
        overview_parts.append(primary_attachment_text[:1200])
    for txts in supporting_texts.values():
        for txt in txts[:2]:
            overview_parts.append(txt[:500])
    overview = "\n".join(p for p in overview_parts if p)

    problem_texts: list[str] = []
    solution_texts: list[str] = []
    value_texts: list[str] = []
    attachment_texts: list[str] = []

    for chunk in chunks:
        text = chunk.get("text", "")
        if not text or chunk.get("is_boilerplate"):
            continue

        probe = f"{chunk.get('section_title', '')} {text[:250]}"
        if chunk.get("attachment_id"):
            attachment_texts.append(text)
        if _PROBLEM_PATTERNS.search(probe):
            problem_texts.append(text)
        elif _SOLUTION_PATTERNS.search(probe):
            solution_texts.append(text)
        elif _VALUE_PATTERNS.search(probe):
            value_texts.append(text)

    if not problem_texts and description_cleaned:
        problem_texts.append(description_cleaned[:900])
    if not solution_texts and description_cleaned:
        solution_texts.append(description_cleaned[:900])
    for comment in comments_cleaned[:2]:
        if _PROBLEM_PATTERNS.search(comment[:200]):
            problem_texts.append(comment[:500])
        elif _SOLUTION_PATTERNS.search(comment[:200]):
            solution_texts.append(comment[:500])

    def _join_limited(texts: list[str], limit: int = 2000) -> str:
        joined = "\n".join(t for t in texts if t)
        return joined[:limit] if joined else ""

    return [
        {"overview": overview[:2200]},
        {"problem_objective": _join_limited(problem_texts, 1800)},
        {"solution_capability": _join_limited(solution_texts, 1800)},
        {"value_proposition": _join_limited(value_texts, 1800)},
        {"attachment_focused": _join_limited(attachment_texts, 2200)},
    ]

def _build_retrieval_text(
    summary_str: str,
    description_cleaned: str,
    primary_attachment_text: str,
    supporting_texts: dict[str, list[str]],
    retrieval_views: dict,
) -> str:
    """Build a single combined retrieval string for ticket-level embedding."""
    parts: list[str] = []
    if summary_str:
        parts.append(summary_str)
    if primary_attachment_text:
        parts.append(primary_attachment_text[:2500])
    elif description_cleaned:
        parts.append(description_cleaned[:1500])

    for key in ("overview", "problem_objective", "solution_capability", "value_proposition"):
        view = retrieval_views.get(key, "")
        if view and view not in "\n".join(parts):
            parts.append(view[:900])

    supporting_concat: list[str] = []
    for role in ("supporting_doc", "primary_idea_card", "primary_fallback"):
        for txt in supporting_texts.get(role, [])[:2]:
            if txt and txt != primary_attachment_text:
                supporting_concat.append(txt[:700])
    if supporting_concat:
        parts.append("\n".join(supporting_concat)[:3000])

    return "\n".join(p.strip() for p in parts if p.strip())[:5000]

def _build_retrieval_chunks(
    raw_chunks: list[dict],
    section_chunks: list[dict],
    section_only_chunks: bool,
    include_section_rollups: bool,
) -> list[dict]:
    """
    Balanced retrieval chunk construction.

    Default behavior is fine-grained chunks only. Section rollups are optional,
    because blindly mixing them with raw chunks over-amplifies primary doc content.
    """
    if section_only_chunks:
        return section_chunks or raw_chunks
    if include_section_rollups:
        return raw_chunks + section_chunks
    return raw_chunks

def _split_oversized_chunks(
    chunks: list[dict],
    max_tokens: int,
    overlap_tokens: int,
) -> list[dict]:
    if max_tokens <= 0:
        return chunks

    overlap_tokens = max(0, min(overlap_tokens, max_tokens // 2))
    step = max(1, max_tokens - overlap_tokens)

    out: list[dict] = []
    split_count = 0
    for chunk in chunks:
        text = str(chunk.get("text") or "")
        words = text.split()
        if len(words) <= max_tokens:
            out.append(chunk)
            continue

        for part, start in enumerate(range(0, len(words), step), start=1):
            piece = words[start : start + max_tokens]
            if not piece:
                break
            new_chunk = dict(chunk)
            new_chunk["text"] = " ".join(piece)
            new_chunk["word_count"] = len(piece)
            new_chunk["chunk_id"] = f"{chunk.get('chunk_id', 'chunk')}_part_{part}"
            # Split chunks should not keep original embeddings
            new_chunk.pop("embedding", None)
            new_chunk["chunk_granularity"] = "split_chunk"
            out.append(new_chunk)
            if start + max_tokens >= len(words):
                break
        split_count += 1

    if split_count:
        logger.info(
            "Split %d oversized retrieval chunks (max_tokens=%d, overlap_tokens=%d)",
            split_count,
            max_tokens,
            overlap_tokens,
        )
    return out

def _doc_role_weight(doc_role: str) -> float:
    if doc_role == "primary_idea_card":
        return 1.00
    if doc_role == "primary_fallback":
        return 0.90
    if doc_role == "supporting_doc":
        return 0.72
    return 0.50

def _derive_content_source(processed_attachments: list[dict], desc_class: str, comments_enriched: dict) -> str:
    if processed_attachments:
        if len(processed_attachments) == 1:
            return "single_attachment"
        return "multi_attachment"
    if desc_class in ("rich", "usable"):
        return "description"
    if comments_enriched.get("comments_cleaned"):
        return "comment"
    return "none"

def _determine_quality_tier_safe(
    content_source: str,
    desc_class: str,
    chunks: list[dict],
    triage_artifact: dict,
) -> str:
    """Local quality tier fallback that does not assume the external signature."""
    attachment_count = len(triage_artifact.get("chunk_candidates", []))
    doc_chunk_count = sum(1 for c in chunks if c.get("attachment_id"))
    avg_conf = sum(float(c.get("extraction_confidence", 0.0) or 0.0) for c in chunks) / max(len(chunks), 1)

    if attachment_count >= 2 and doc_chunk_count >= 8 and avg_conf >= 0.85:
        return "A"
    if (attachment_count >= 1 and doc_chunk_count >= 3) or (desc_class in ("rich", "usable") and len(chunks) >= 3):
        return "B"
    if content_source != "none" and chunks:
        return "C"
    return "D"

def _heuristic_summary(meta: dict, description_cleaned: str, processed_attachments: list[dict]) -> str:
    parts = [meta.get("summary", "")]
    if description_cleaned:
        parts.append(description_cleaned[:800])
    for doc in processed_attachments[:2]:
        if doc.get("preview"):
            parts.append(doc["preview"])
    return "\n".join(p for p in parts if p).strip()[:1000]

def _source_quality_score(quality_tier: str, avg_confidence: float, processed_attachments: list[dict]) -> float:
    tier_weights = {"A": 1.00, "B": 0.82, "C": 0.65, "D": 0.40}
    attachment_bonus = min(0.12, 0.03 * sum(1 for a in processed_attachments if a.get("extraction_status") == "extracted"))
    return tier_weights.get(quality_tier, 0.40) * 0.80 + avg_confidence * 0.20 + attachment_bonus

def _build_ticket_source_url(ticket_data: dict, ticket_key: str) -> str:
    issue_self = ticket_data.get("self", "")
    if isinstance(issue_self, str) and issue_self:
        return issue_self
    return ticket_key

def _stable_chunk_uid(
    ticket_key: str,
    attachment_id: str,
    source: str,
    chunk_id: str,
) -> str:
    raw = f"{ticket_key}::{attachment_id}::{source}::{chunk_id}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:32]

def _clean_chunk_text(text: str) -> str:
    """Normalize whitespace and encoding artifacts in chunk text."""
    if not text:
        return ""
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    return text.strip()

def _chunk_source_locator(chunk: dict) -> str:
    if chunk.get("page_num") is not None:
        return f"Page {chunk.get('page_num')}"
    if chunk.get("slide_num") is not None:
        return f"Slide {chunk.get('slide_num')}"
    if chunk.get("page_range"):
        start, end = chunk.get("page_range", (None, None))
        return f"Pages {start}-{end}"
    if chunk.get("slide_range"):
        start, end = chunk.get("slide_range", (None, None))
        return f"Slides {start}-{end}"
    return str(chunk.get("chunk_id", ""))

def _chunk_header_hierarchy(chunk: dict) -> str:
    parts: list[str] = []
    section_title = str(chunk.get("section_title") or "").strip()
    slide_title = str(chunk.get("slide_title") or "").strip()
    if section_title:
        parts.append(section_title)
    if slide_title and slide_title != section_title:
        parts.append(slide_title)
    locator = _chunk_source_locator(chunk)
    if locator:
        parts.append(locator)
    return " > ".join(parts)

def _source_for_ext(ext: str) -> str:
    ext = ext.lower()
    if ext in ("pptx", "ppt"):
        return "pptx_slide"
    if ext == "pdf":
        return "pdf_page"
    if ext in ("docx", "doc"):
        return "doc_section"
    if ext in ("xlsx", "xls"):
        return "sheet_row"
    if ext == "csv":
        return "csv_row"
    if ext in ("png", "jpg", "jpeg", "gif", "bmp", "tiff", "webp"):
        return "image_ocr"
    return "attachment"

def _granularity_for_source(source: str) -> str:
    if source in ("pptx_slide", "pdf_page"):
        return "page_or_slide"
    if source in ("doc_section", "sheet_row", "csv_row", "image_ocr"):
        return "attachment_fragment"
    if source == "section":
        return "section_rollup"
    return "chunk"

def _extract_method_for_ext(ext: str) -> str:
    ext = ext.lower()
    if ext in ("pptx", "ppt", "pdf", "docx", "doc"):
        return "native"
    if ext in ("xlsx", "xls", "csv"):
        return "tabular"
    if ext in ("png", "jpg", "jpeg", "gif", "bmp", "tiff", "webp"):
        return "ocr"
    return "none"

def _ext_from_name(filename: str) -> str:
    filename = str(filename or "")
    return filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

def _attach_chunk_identity(
    chunks: list[dict],
    ticket_key: str,
    source_url: str,
    default_attachment_id: str = "",
    default_attachment_name: str = "",
) -> None:
    for idx, chunk in enumerate(chunks):
        chunk_id = str(chunk.get("chunk_id") or f"chunk-{idx}")
        source = str(chunk.get("source") or "unknown")

        if chunk.get("no_attachment"):
            attachment_id = ""
            attachment_name = ""
        else:
            attachment_id = str(chunk.get("attachment_id") or default_attachment_id)
            attachment_name = str(chunk.get("attachment_name") or default_attachment_name)

        chunk["chunk_index"] = idx
        chunk["token_count"] = int(chunk.get("word_count") or len(str(chunk.get("text", "")).split()))
        chunk["source_url"] = ticket_key
        chunk["source_uri"] = attachment_id
        chunk["attachment_id"] = attachment_id
        chunk["attachment_name"] = attachment_name
        chunk["attachment_type"] = _ext_from_name(attachment_name) if attachment_name else ""
        chunk["header_hierarchy"] = _chunk_header_hierarchy(chunk)
        chunk["chunk_uid"] = _stable_chunk_uid(
            ticket_key=ticket_key,
            attachment_id=attachment_id,
            source=source,
            chunk_id=chunk_id,
        )
        chunk.pop("no_attachment", None)
        if chunk.get("text"):
            chunk["text"] = _clean_chunk_text(chunk["text"])